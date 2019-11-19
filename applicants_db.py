import os
import glob
import logging
import hashlib
import unicodedata

from enum import Enum

import openpyxl


class XLSXHeaders(Enum):
    position = 0
    fio = 1
    money = 2
    comment = 3
    status = 4


class ApplicantsDBError(Exception):
    pass


class Applicant:
    def __init__(self, position, fio, money, comment, status, resume_file_tmpl, uuid):
        """
        :param position:
        :param fio:
        :param money:
        :param comment:
        :param status:
        :param resume_file_tmpl: like '/path_to_db/position/fio.*'
        """
        self.position = position
        self.fio = fio
        self.money = money
        self.comment = comment
        self.status = status
        self.resume_file_tmpl = resume_file_tmpl
        self.uuid = uuid

    @property
    def resume_file(self):
        # TODO: cache
        files = glob.glob(unicodedata.normalize('NFD', self.resume_file_tmpl))
        if len(files) != 1:
            msg = f'can not choose resume file from "{files}" (by "{self.resume_file_tmpl}")'
            logging.error(msg)
            raise ApplicantsDBError(msg)
        return files[0]

    def __str__(self):
        return f'{self.position}, {self.fio}, {self.money} ({self.uuid})'

    def __repr__(self):
        return (f'{self.__class__.__name__}("{self.position}", "{self.fio}", "{self.money}", '
                f'"{self.comment}", "{self.status}", "{self.resume_file_tmpl}", "{self.uuid}")')


class ApplicantsDBXLSX:
    def __init__(self, db_dir, xlsx_filename=None):
        self.db_dir = os.path.abspath(db_dir)
        self.xlsx_path = self.get_xlsx_path(xlsx_filename)

    def get_xlsx_path(self, xlsx_filename):
        if xlsx_filename:
            full_name = os.path.join(self.db_dir, xlsx_filename)
            if not os.path.exists(full_name):
                msg = f'xlsx file of aplicants db "{full_name}" is not exists'
                logging.error(msg)
                raise ApplicantsDBError(msg)
            return full_name

        files = glob.glob(os.path.join(self.db_dir, '*.xlsx'))
        if not files:
            msg = f'database directory "{self.db_dir}" has not *.xlsx file'
            logging.error(msg)
            raise ApplicantsDBError(msg)

        if len(files) != 1:
            msg = f'database directory "{self.db_dir}" has several *.xlsx - please specify which one'
            logging.error(msg)
            raise ApplicantsDBError(msg)

        return files[0]

    def __str__(self):
        return self.xlsx_path

    @staticmethod
    def _get_cell_value(row, enum_field):
        return row[enum_field.value].value.strip()

    @staticmethod
    def _get_money_value(row, enum_field):
        value = row[enum_field.value].value
        if isinstance(value, str):
            return value.strip()
        # note: is float valid?
        return str(int(value))

    def _get_resume_file_tmpl(self, position, fio):
        return os.path.join(self.db_dir, position, f'{fio}.*')

    @staticmethod
    def _get_applicant_uuid(position, fio, money):
        return hashlib.md5(f'{position}{fio}{money}'.encode('utf-8')).hexdigest()

    def get_applicants(self):
        book = openpyxl.load_workbook(self.xlsx_path)
        sheets = book.sheetnames
        logging.debug(f'"{self.xlsx_path}" has "{sheets}" sheets')

        for sheet_name in sheets:
            sheet = book[sheet_name]
            for row in sheet.iter_rows(min_row=2):
                position = self._get_cell_value(row, XLSXHeaders.position)
                fio = self._get_cell_value(row, XLSXHeaders.fio)
                money = self._get_money_value(row, XLSXHeaders.money)
                yield Applicant(position=position, fio=fio, money=money,
                                comment=self._get_cell_value(row, XLSXHeaders.comment),
                                status=self._get_cell_value(row, XLSXHeaders.status),
                                resume_file_tmpl=self._get_resume_file_tmpl(position, fio),
                                uuid=self._get_applicant_uuid(position, fio, money))
